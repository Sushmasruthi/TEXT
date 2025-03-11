from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    jsonify,
    send_file,
)
from database import (
    init_db,
    register_student,
    register_teacher,
    verify_student,
    verify_teacher,
    Database,
    ResultsDatabase,
)
from functools import wraps
import os
from werkzeug.utils import secure_filename
from image_to_text import extract_text_from_image
from text_to_json import process_text_with_image
import pandas as pd
import json
from PIL import Image
from datetime import datetime
import re
import shutil
from zipfile import ZipFile
import tempfile
import openpyxl
import time
import threading

app = Flask(__name__)
app.secret_key = "your_secret_key_here"  # Change this to a secure secret key

# Initialize the database
init_db()

db = Database()
db_results = ResultsDatabase()

# Add these configurations
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Create uploads directory if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

TEMP_FOLDER = "temp"
os.makedirs(TEMP_FOLDER, exist_ok=True)


# Helper function to check allowed file extensions
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_type" not in session:
            flash("Please log in first", "error")
            return redirect(url_for("index"))
        return f(*args, **kwargs)

    return decorated_function


@app.route("/")
def index():
    return render_template("portal.html")


@app.route("/student/register", methods=["GET", "POST"])
def student_register():
    if request.method == "POST":
        try:
            # Try to get JSON data first
            if request.is_json:
                data = request.get_json()
                student_id = data.get("studentId")
                full_name = data.get("fullName")
                department = data.get("department")
                password = data.get("password")
            else:
                # Fall back to form data
                student_id = request.form.get("studentId")
                full_name = request.form.get("fullName")
                department = request.form.get("department")
                password = request.form.get("password")

            if not all([student_id, full_name, department, password]):
                return (
                    jsonify({"success": False, "message": "All fields are required"}),
                    400,
                )

            success, message = register_student(
                student_id, full_name, department, password
            )

            if success:
                return (
                    jsonify(
                        {
                            "success": True,
                            "message": "Registration successful! Please login.",
                        }
                    ),
                    200,
                )
            else:
                return jsonify({"success": False, "message": message}), 400

        except Exception as e:
            return jsonify({"success": False, "message": str(e)}), 500

    return render_template("student_register.html")


@app.route("/teacher/register", methods=["GET", "POST"])
def teacher_register():
    if request.method == "POST":
        try:
            # Try to get JSON data first
            if request.is_json:
                data = request.get_json()
                teacher_id = data.get("teacherId")
                full_name = data.get("fullName")
                department = data.get("department")
                specialization = data.get("specialization")
                password = data.get("password")
            else:
                # Fall back to form data
                teacher_id = request.form.get("teacherId")
                full_name = request.form.get("fullName")
                department = request.form.get("department")
                specialization = request.form.get("specialization")
                password = request.form.get("password")

            if not all([teacher_id, full_name, department, specialization, password]):
                return (
                    jsonify({"success": False, "message": "All fields are required"}),
                    400,
                )

            success, message = register_teacher(
                teacher_id, full_name, department, specialization, password
            )

            if success:
                return (
                    jsonify(
                        {
                            "success": True,
                            "message": "Registration successful! Please login.",
                        }
                    ),
                    200,
                )
            else:
                return jsonify({"success": False, "message": message}), 400

        except Exception as e:
            return jsonify({"success": False, "message": str(e)}), 500

    return render_template("teacher_register.html")


@app.route("/student/login", methods=["GET", "POST"])
def student_login():
    if request.method == "POST":
        try:
            # Try to get JSON data first
            if request.is_json:
                data = request.get_json()
                student_id = data.get("studentId")
                password = data.get("password")
            else:
                # Fall back to form data
                student_id = request.form.get("studentId")
                password = request.form.get("password")

            if not student_id or not password:
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": "Both ID and password are required",
                        }
                    ),
                    400,
                )

            success, result = verify_student(student_id, password)

            if success:
                session["user_id"] = student_id
                session["user_type"] = "student"
                session["full_name"] = result["full_name"]
                return jsonify({"success": True, "message": "Login successful"})
            else:
                return jsonify({"success": False, "message": result}), 401

        except Exception as e:
            return jsonify({"success": False, "message": str(e)}), 500

    return render_template("student_login.html")


@app.route("/teacher/login", methods=["GET", "POST"])
def teacher_login():
    if request.method == "POST":
        try:
            # Try to get JSON data first
            if request.is_json:
                data = request.get_json()
                teacher_id = data.get("teacherId")
                password = data.get("password")
            else:
                # Fall back to form data
                teacher_id = request.form.get("teacherId")
                password = request.form.get("password")

            if not teacher_id or not password:
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": "Both ID and password are required",
                        }
                    ),
                    400,
                )

            success, result = verify_teacher(teacher_id, password)

            if success:
                session["user_id"] = teacher_id
                session["user_type"] = "teacher"
                session["full_name"] = result["full_name"]
                return jsonify({"success": True, "message": "Login successful"})
            else:
                return jsonify({"success": False, "message": result}), 401

        except Exception as e:
            return jsonify({"success": False, "message": str(e)}), 500

    return render_template("teacher_login.html")


@app.route("/student/dashboard")
@login_required
def student_dashboard():
    if session.get("user_type") != "student":
        flash("Unauthorized access", "error")
        return redirect(url_for("index"))
    return render_template("student_dashboard.html", name=session.get("full_name"))


@app.route("/teacher/dashboard")
@login_required
def teacher_dashboard():
    if session.get("user_type") != "teacher":
        flash("Unauthorized access", "error")
        return redirect(url_for("index"))
    return render_template(
        "teacher_dashboard.html",
        teacher_name=session.get("full_name"),
        department=session.get("department"),
    )


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out successfully.", "success")
    return redirect(url_for("index"))


@app.route("/upload")
@login_required
def upload_page():
    if session.get("user_type") != "teacher":
        flash("Unauthorized access", "error")
        return redirect(url_for("index"))
    return render_template("upload.html")


@app.route("/api/analysis")
@login_required
def get_analysis():
    year = request.args.get("year")
    subject = request.args.get("subject")
    exam_type = request.args.get("examType")

    if not all([year, subject, exam_type]):
        return jsonify({"error": "Missing required parameters"}), 400

    try:
        analysis = db_results.get_detailed_analysis(year, subject, exam_type)
        return jsonify(analysis)
    except Exception as e:
        print(f"Analysis error: {str(e)}")
        return jsonify({"error": "Failed to fetch analysis"}), 500


def process_files(files):
    results = []
    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            file.save(filepath)

            try:
                # Extract text from image
                extracted_result = extract_text_from_image(filepath)
                print(f"Extracted result: {extracted_result}")

                # Check if extraction was successful
                if not extracted_result or not isinstance(extracted_result, dict):
                    print(f"Invalid extraction result from {filename}")
                    continue

                # Get the extracted text from the result
                extracted_text = extracted_result.get("text")
                if not extracted_text:
                    print(f"No text content extracted from {filename}")
                    continue

                # Process the extracted text into structured data
                processed_data = process_text_with_image(extracted_text, filepath)

                if not processed_data:
                    print(f"Failed to process data from {filename}")
                    continue

                # Validate the processed data structure
                if not isinstance(processed_data, dict):
                    print(f"Invalid data format from {filename}")
                    continue

                # Ensure required fields exist
                required_fields = ["roll_number", "questions", "total_marks"]
                if not all(field in processed_data for field in required_fields):
                    print(f"Missing required fields in data from {filename}")
                    continue

                # Validate questions structure
                questions = processed_data.get("questions", {})
                if not isinstance(questions, dict):
                    print(f"Invalid questions format in {filename}")
                    continue

                # Validate question data structure
                valid_question = True
                for q_num in range(1, 7):
                    q_key = f"Q{q_num}"
                    if q_key not in questions:
                        questions[q_key] = {"a": 0, "b": 0, "c": 0, "d": 0}
                        continue

                    q_data = questions[q_key]
                    if not isinstance(q_data, dict) or not all(
                        part in q_data for part in ["a", "b", "c", "d"]
                    ):
                        print(f"Invalid format for question {q_key} in {filename}")
                        valid_question = False
                        break

                if not valid_question:
                    continue

                # Add to results if all validation passes
                results.append(processed_data)

            except Exception as e:
                print(f"Error processing {filename}: {str(e)}")
            finally:
                # Clean up temporary file
                if os.path.exists(filepath):
                    os.remove(filepath)

    return results


@app.route("/upload-folder", methods=["POST"])
@login_required
def upload_folder():
    if session.get("user_type") != "teacher":
        return jsonify({"success": False, "message": "Unauthorized access"}), 403

    try:
        class_year = request.form.get("class")
        subject = request.form.get("subject")
        exam_type = request.form.get("examType")
        academic_year = str(datetime.now().year)

        if not all([class_year, subject, exam_type]):
            return (
                jsonify({"success": False, "message": "Missing required fields"}),
                400,
            )

        results = []
        files_to_process = []

        # Handle individual file uploads
        if "files[]" in request.files:
            files = request.files.getlist("files[]")
            if files and files[0].filename:
                files_to_process.extend([f for f in files if allowed_file(f.filename)])

        # Handle folder upload
        if "folder[]" in request.files:
            folder_files = request.files.getlist("folder[]")
            if folder_files and folder_files[0].filename:
                files_to_process.extend(
                    [f for f in folder_files if allowed_file(f.filename)]
                )

        if not files_to_process:
            return (
                jsonify({"success": False, "message": "No valid image files found"}),
                400,
            )

        # Process all collected files
        for file in files_to_process:
            try:
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(filepath)

                result = process_single_image(filepath)
                if result:
                    results.append(result)

                # Clean up the temporary file
                if os.path.exists(filepath):
                    os.remove(filepath)

            except Exception as e:
                print(f"Error processing {file.filename}: {str(e)}")
                continue

        if not results:
            return (
                jsonify({"success": False, "message": "No valid data extracted"}),
                400,
            )

        # Save to database
        db_results.save_results(results, class_year, subject, exam_type, academic_year)

        # Store in session for display
        session["upload_results"] = results

        return jsonify(
            {
                "success": True,
                "message": f"Successfully processed {len(results)} files",
                "redirect": url_for("show_results"),
            }
        )

    except Exception as e:
        print(f"Upload error: {str(e)}")
        return (
            jsonify({"success": False, "message": f"Error processing files: {str(e)}"}),
            500,
        )


def process_single_image(file_path):
    """Process a single image file and return extracted data."""
    try:
        # Extract text from image
        extracted_result = extract_text_from_image(file_path)

        if not extracted_result or not isinstance(extracted_result, dict):
            print(f"Invalid extraction result from {file_path}")
            return None

        # Get the extracted text from the result
        extracted_text = extracted_result.get("text")
        if not extracted_text:
            print(f"No text content extracted from {file_path}")
            return None

        # Process the extracted text into structured data
        processed_data = process_text_with_image(extracted_text, file_path)

        if not processed_data:
            print(f"Failed to process data from {file_path}")
            return None

        return validate_processed_data(processed_data)

    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        return None


def validate_processed_data(data):
    """Validate processed data structure."""
    if not isinstance(data, dict):
        return None

    # Ensure required fields exist
    required_fields = ["roll_number", "questions", "total_marks"]
    if not all(field in data for field in required_fields):
        return None

    # Validate questions structure
    questions = data.get("questions", {})
    if not isinstance(questions, dict):
        return None

    # Validate question data structure
    for q_num in range(1, 7):
        q_key = f"Q{q_num}"
        if q_key not in questions:
            questions[q_key] = {"a": 0, "b": 0, "c": 0, "d": 0}
            continue

        q_data = questions[q_key]
        if not isinstance(q_data, dict) or not all(
            part in q_data for part in ["a", "b", "c", "d"]
        ):
            return None

        # Validate mark values
        for part in ["a", "b", "c", "d"]:
            mark = q_data[part]
            if not isinstance(mark, (int, float)) or not (0 <= mark <= 8):
                q_data[part] = 0

    return data


def extract_roll_number(text):
    """Extract roll number from text"""
    match = re.search(r"Roll No:?\s*([A-Z0-9]+)", text)
    return match.group(1) if match else "000000000000"


def extract_question_marks(text):
    """Extract question marks from text"""
    questions = {}
    for i in range(1, 7):
        questions[f"Q{i}"] = {"a": 0, "b": 0, "c": 0, "d": 0}

    # Extract marks using regex
    pattern = r"Q(\d+)[:\s]+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)"
    matches = re.finditer(pattern, text)

    for match in matches:
        q_num = match.group(1)
        if 1 <= int(q_num) <= 6:
            questions[f"Q{q_num}"] = {
                "a": float(match.group(2)),
                "b": float(match.group(3)),
                "c": float(match.group(4)),
                "d": float(match.group(5)),
            }

    return questions


def calculate_total_marks(text):
    """Calculate total marks from text"""
    match = re.search(r"Total[:\s]+(\d+)", text)
    return float(match.group(1)) if match else 0


def get_or_create_subject(subject_name, class_name, academic_year):
    """Helper function to get or create subject ID"""
    with db.get_connection() as conn:
        cursor = conn.cursor()

        # Get or create class
        cursor.execute(
            """
            SELECT id FROM classes 
            WHERE year = ? AND academic_year = ?
        """,
            (class_name, academic_year),
        )

        class_result = cursor.fetchone()
        if class_result:
            class_id = class_result[0]
        else:
            cursor.execute(
                """
                INSERT INTO classes (year, department, academic_year)
                VALUES (?, ?, ?)
            """,
                (class_name, "DEFAULT", academic_year),
            )
            class_id = cursor.lastrowid

        # Get or create subject
        cursor.execute(
            """
            SELECT id FROM subjects 
            WHERE name = ? AND class_id = ?
        """,
            (subject_name, class_id),
        )

        subject_result = cursor.fetchone()
        if subject_result:
            return subject_result[0]

        cursor.execute(
            """
            INSERT INTO subjects (name, class_id, teacher_id)
            VALUES (?, ?, ?)
        """,
            (subject_name, class_id, session.get("user_id")),
        )

        return cursor.lastrowid


@app.route("/results")
@login_required
def show_results():
    if session.get("user_type") != "teacher":
        flash("Unauthorized access", "error")
        return redirect(url_for("index"))

    # Get results from session
    json_data = session.get("upload_results", [])

    return render_template("results.html", json_data=json_data)


@app.route("/delete_last", methods=["POST"])
@login_required
def delete_last():
    if session.get("user_type") != "teacher":
        return jsonify({"success": False, "message": "Unauthorized access"}), 403

    json_data = session.get("upload_results", [])

    if not json_data:
        return jsonify({"success": False, "message": "No entries to delete"}), 400

    # Remove the last entry
    json_data.pop()
    session["upload_results"] = json_data

    return jsonify({"success": True, "message": "Last entry deleted successfully"})


@app.route("/download_excel")
@login_required
def download_excel():
    if session.get("user_type") != "teacher":
        flash("Unauthorized access", "error")
        return redirect(url_for("index"))

    json_data = session.get("upload_results", [])

    if not json_data:
        flash("No data available to download", "error")
        return redirect(url_for("show_results"))

    # Create a DataFrame from the JSON data
    rows = []
    for entry in json_data:
        row = {
            "Roll Number": entry.get("roll_number", ""),
        }

        # Add question marks
        for q_num in range(1, 7):
            q_key = f"Q{q_num}"
            if q_key in entry.get("questions", {}):
                q_data = entry["questions"][q_key]
                row[f"{q_key}a"] = q_data.get("a", 0)
                row[f"{q_key}b"] = q_data.get("b", 0)
                row[f"{q_key}c"] = q_data.get("c", 0)
                row[f"{q_key}d"] = q_data.get("d", 0)
            else:
                row[f"{q_key}a"] = 0
                row[f"{q_key}b"] = 0
                row[f"{q_key}c"] = 0
                row[f"{q_key}d"] = 0

        row["Total"] = entry.get("total_marks", 0)
        rows.append(row)

    df = pd.DataFrame(rows)

    # Create Excel file
    excel_file = os.path.join(app.config["UPLOAD_FOLDER"], "results.xlsx")
    df.to_excel(excel_file, index=False)

    # Return the file for download
    return send_file(excel_file, as_attachment=True)


@app.route("/marks-analysis")
@login_required
def marks_analysis():
    if session.get("user_type") != "teacher":
        flash("Unauthorized access", "error")
        return redirect(url_for("index"))

    with db.get_connection() as conn:
        cursor = conn.cursor()

        # Get all classes from the classes table
        cursor.execute(
            """
            SELECT DISTINCT c.year, c.academic_year 
            FROM classes c
            ORDER BY c.academic_year DESC, c.year
        """
        )
        class_years = [f"{row[0]} ({row[1]})" for row in cursor.fetchall()]

        # Get all subjects from the subjects table
        cursor.execute(
            """
            SELECT DISTINCT s.name 
            FROM subjects s
            JOIN classes c ON s.class_id = c.id
            ORDER BY s.name
        """
        )
        all_subjects = [row[0] for row in cursor.fetchall()]

    with db_results.get_connection() as conn:
        cursor = conn.cursor()

        # Get all unique class years from results
        cursor.execute(
            """
            SELECT DISTINCT class_year 
            FROM students_results 
            ORDER BY class_year
        """
        )
        result_years = [row[0] for row in cursor.fetchall()]

        # Get all unique subjects from results
        cursor.execute(
            """
            SELECT DISTINCT subject 
            FROM students_results 
            ORDER BY subject
        """
        )
        result_subjects = [row[0] for row in cursor.fetchall()]

        # Get all unique exam types
        cursor.execute(
            """
            SELECT DISTINCT exam_type 
            FROM students_results 
            ORDER BY exam_type
        """
        )
        exam_types = [row[0] for row in cursor.fetchall()]

        # Get summary statistics for each class and subject
        cursor.execute(
            """
            SELECT 
                class_year,
                subject,
                exam_type,
                COUNT(*) as total_students,
                AVG(total_marks) as avg_marks,
                MAX(total_marks) as max_marks,
                MIN(total_marks) as min_marks,
                COUNT(CASE WHEN total_marks >= 20 THEN 1 END) as passed_count
            FROM students_results
            GROUP BY class_year, subject, exam_type
            ORDER BY class_year, subject, exam_type
        """
        )

        class_stats = {}
        for row in cursor.fetchall():
            if row[0] not in class_stats:
                class_stats[row[0]] = {}
            if row[1] not in class_stats[row[0]]:
                class_stats[row[0]][row[1]] = {}

            total_students = row[3]
            class_stats[row[0]][row[1]][row[2]] = {
                "total_students": total_students,
                "avg_marks": round(row[4], 2) if row[4] else 0,
                "max_marks": row[5],
                "min_marks": row[6],
                "pass_percentage": (
                    round((row[7] / total_students * 100), 2)
                    if total_students > 0
                    else 0
                ),
            }

    # Combine years and subjects from both databases
    years = sorted(set(class_years + result_years))
    subjects = sorted(set(all_subjects + result_subjects))

    return render_template(
        "marks_analysis.html",
        teacher_name=session.get("full_name"),
        years=years,
        subjects=subjects,
        exam_types=exam_types,
        class_stats=class_stats,
    )


@app.route("/view-marks")
@login_required
def view_marks():
    if session.get("user_type") != "teacher":
        flash("Unauthorized access", "error")
        return redirect(url_for("index"))

    # Get all years and exam types for the dropdowns
    with db_results.get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT DISTINCT class_year FROM students_results ORDER BY class_year"
        )
        years = [row[0] for row in cursor.fetchall()]

        cursor.execute(
            "SELECT DISTINCT exam_type FROM students_results ORDER BY exam_type"
        )
        exam_types = [row[0] for row in cursor.fetchall()]

        print(f"Years: {years}")
        print(f"Exam Types: {exam_types}")

    return render_template("view_marks.html", years=years, exam_types=exam_types)


@app.route("/api/view-marks")
@login_required
def get_marks():
    year = request.args.get("year")
    subject = request.args.get("subject")
    exam_type = request.args.get("examType")

    with db_results.get_connection() as conn:
        cursor = conn.cursor()

        # Get marks with question details
        cursor.execute(
            """
            SELECT sr.roll_number, sr.subject, sr.total_marks,
                   qm.question_number, qm.part_a, qm.part_b, qm.part_c, qm.part_d
            FROM students_results sr
            LEFT JOIN question_marks qm ON sr.id = qm.result_id
            WHERE sr.class_year = ? AND sr.subject = ? AND sr.exam_type = ?
            ORDER BY sr.roll_number, qm.question_number
        """,
            (year, subject, exam_type),
        )

        results = {}
        for row in cursor.fetchall():
            roll_number = row[0]
            if roll_number not in results:
                results[roll_number] = {
                    "roll_number": roll_number,
                    "subject": row[1],
                    "total_marks": row[2],
                    "questions": {},
                }

            if row[3]:  # if there are question details
                q_num = row[3]
                results[roll_number]["questions"][f"Q{q_num}"] = {
                    "a": row[4],
                    "b": row[5],
                    "c": row[6],
                    "d": row[7],
                }

        return jsonify(list(results.values()))


@app.route("/api/update-marks", methods=["POST"])
@login_required
def update_marks():
    if session.get("user_type") != "teacher":
        return jsonify({"success": False, "message": "Unauthorized access"}), 403

    try:
        data = request.get_json()
        roll_number = data.get("roll_number")
        class_year = data.get("class_year")
        subject = data.get("subject")
        exam_type = data.get("exam_type")
        question_marks = data.get("questions")
        total_marks = data.get("total_marks")

        if not all(
            [roll_number, class_year, subject, exam_type, question_marks, total_marks]
        ):
            return (
                jsonify({"success": False, "message": "Missing required fields"}),
                400,
            )

        success, message = db_results.update_result(
            roll_number, class_year, subject, exam_type, question_marks, total_marks
        )

        return jsonify({"success": success, "message": message})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/delete-marks", methods=["POST"])
@login_required
def delete_marks():
    if session.get("user_type") != "teacher":
        return jsonify({"success": False, "message": "Unauthorized access"}), 403

    try:
        data = request.get_json()
        roll_number = data.get("roll_number")
        class_year = data.get("class_year")
        subject = data.get("subject")
        exam_type = data.get("exam_type")

        if not all([roll_number, class_year, subject, exam_type]):
            return (
                jsonify({"success": False, "message": "Missing required fields"}),
                400,
            )

        success, message = db_results.delete_result(
            roll_number, class_year, subject, exam_type
        )

        return jsonify({"success": success, "message": message})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/download-marks")
@login_required
def download_marks():
    if session.get("user_type") != "teacher":
        flash("Unauthorized access", "error")
        return redirect(url_for("index"))

    year = request.args.get("year")
    subject = request.args.get("subject")
    exam_type = request.args.get("examType")

    if not all([year, subject, exam_type]):
        flash("Missing required parameters", "error")
        return redirect(url_for("view_marks"))

    with db_results.get_connection() as conn:
        cursor = conn.cursor()

        # Get marks with question details
        cursor.execute(
            """
            SELECT sr.roll_number, sr.subject, sr.total_marks,
                   qm.question_number, qm.part_a, qm.part_b, qm.part_c, qm.part_d
            FROM students_results sr
            LEFT JOIN question_marks qm ON sr.id = qm.result_id
            WHERE sr.class_year = ? AND sr.subject = ? AND sr.exam_type = ?
            ORDER BY sr.roll_number, qm.question_number
        """,
            (year, subject, exam_type),
        )

        results = {}
        for row in cursor.fetchall():
            roll_number = row[0]
            if roll_number not in results:
                results[roll_number] = {
                    "Roll Number": roll_number,
                    "Subject": row[1],
                    "Total Marks": row[2],
                }
                # Initialize all question parts with 0
                for q_num in range(1, 7):
                    results[roll_number][f"Q{q_num}a"] = 0
                    results[roll_number][f"Q{q_num}b"] = 0
                    results[roll_number][f"Q{q_num}c"] = 0
                    results[roll_number][f"Q{q_num}d"] = 0

            if row[3]:  # if there are question details
                q_num = row[3]
                results[roll_number][f"Q{q_num}a"] = row[4]
                results[roll_number][f"Q{q_num}b"] = row[5]
                results[roll_number][f"Q{q_num}c"] = row[6]
                results[roll_number][f"Q{q_num}d"] = row[7]

        if not results:
            flash("No data available to download", "error")
            return redirect(url_for("view_marks"))

        # Create DataFrame and Excel file
        df = pd.DataFrame(list(results.values()))

        # Create a temporary file
        temp_dir = tempfile.mkdtemp()
        excel_path = os.path.join(temp_dir, f"marks_{year}_{subject}_{exam_type}.xlsx")

        # Save to Excel with formatting
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Marks")

            # Get the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets["Marks"]

            # Format headers
            for col in range(worksheet.max_column):
                cell = worksheet.cell(row=1, column=col + 1)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="8B5CF6", end_color="8B5CF6", fill_type="solid"
                )
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[
                    openpyxl.utils.get_column_letter(column[0].column)
                ].width = adjusted_width

        # Send file and clean up
        try:
            return send_file(
                excel_path,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"marks_{year}_{subject}_{exam_type}.xlsx",
            )
        finally:
            # Clean up temp file in a separate thread to avoid blocking
            def cleanup():
                time.sleep(1)  # Wait a bit to ensure file is sent
                shutil.rmtree(temp_dir, ignore_errors=True)

            threading.Thread(target=cleanup).start()


if __name__ == "__main__":
    app.run(debug=True)
